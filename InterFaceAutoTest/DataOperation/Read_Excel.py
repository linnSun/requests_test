# -*- coding: utf-8 -*-#
#-------------------------------------------------------------------------------
# ProjectName:     Test0805
# FileName:       Read_Excel 
# Author:        MuTou
# Date:         2019/10/16
# Description:实现获取excel当中的数据信息
#-------------------------------------------------------------------------------
#通过第三方模块：openpyxl
#Workbook表示获取创建一个工作簿的对象，操作一个已存在的工作簿所以需要加载
from openpyxl import load_workbook
from InterFaceAutoTest.DataOperation.ExcelColumn import Excel_Column
from InterFaceAutoTest.DataOperation.Read_Json import ReadJson
class  ReadExcel:
    def __init__(self,ExcelPath,SheetName,ParamJsonPath,ResponseJsonPath):
        #初始化读取json的对象
        self.read_param_json=ReadJson(ParamJsonPath)
        self.read_response_json=ReadJson(ResponseJsonPath)
        #需要获取操作excel的对象
        self.get_excel_object=load_workbook(ExcelPath)
        self.get_write_object=load_workbook(ExcelPath)
        #获取指定操作的sheet
        self.get_Sheet=self.get_excel_object[SheetName]
        self.get_Write_Sheet=self.get_write_object[SheetName]

    #获取单元格对象
    def  get_cell(self,columnName,rowNum):
        return self.get_Sheet[columnName+str(rowNum)]

    def get_write_cell(self, columnName, rowNum):
        return self.get_Write_Sheet[columnName + str(rowNum)]
    #获取动态单元格的值
    def  get_cell_value(self,columnName,rowNum):
        return self.get_cell(columnName,rowNum).value

    #获取url、method、param
    def  get_cell_url_value(self,rowNum):
        return self.get_cell_value(Excel_Column.CASE_URL,rowNum)

    #获取方法：
    def  get_cell_method_value(self,rowNum):
        return self.get_cell_value(Excel_Column.CASE_METHOD,rowNum)
    #获取参数
    def  get_cell_param_value(self,rowNum):
        get_key=self.get_cell_value(Excel_Column.CASE_PARAM,rowNum)
        return self.read_param_json.get_excelkey_value(get_key)

    #获取预期结果值
    def  get_cell_expect_value(self,rowNum):
        get_key = self.get_cell_value(Excel_Column.CASE_EXPECT, rowNum)
        return self.read_response_json.get_excelkey_value(get_key)

    #获取是否通过的单元格的对象
    def  get_cell_result_object(self,rowNum):
        return self.get_write_cell(Excel_Column.CASE_RESULT,rowNum)

    #获取实际结果的单元格对象
    def  get_cell_actual_object(self,rowNum):
        return self.get_write_cell(Excel_Column.CASE_ACTUAL,rowNum)


    #获取是否执行的结果值
    def  get_cell_execute_value(self,rowNum):
        return self.get_cell_value(Excel_Column.CASE_EXECUTE,rowNum)

    #获取当excel当中指定sheet的最大行
    def  get_max_row(self):
        return self.get_Sheet.max_row

    #需要获取是否依赖、依赖业务数据、依赖cookie
    def  get_cell_depend_value(self,rowNum):
        return self.get_cell_value(Excel_Column.CASE_DEPEND,rowNum)

    def get_cell_business_value(self,rowNum):
        return self.get_cell_value(Excel_Column.CASE_BUSINESS,rowNum)

    def get_cell_cookie_value(self, rowNum):
        return self.get_cell_value(Excel_Column.CASE_COOKIE, rowNum)

    #获取excel中的id
    def get_cell_id_value(self, rowNum):
        return self.get_cell_value(Excel_Column.CASE_ID ,rowNum)

    #获取依赖的字段、以及被依赖的字段
    def  get_cell_from_depend_column_value(self,rowNum):
        return self.get_cell_value(Excel_Column.CASE_FROM_COLUMN, rowNum)

    def get_cell_depend_column_value(self, rowNum):
        return self.get_cell_value(Excel_Column.CASE_COLUMN, rowNum)


if __name__ == '__main__':
    rd=ReadExcel("./test0805.xlsx","Sheet1","./ParameterData.json","./ExpectResponse.json")
    for i in range(2,rd.get_max_row()+1):
        # print(rd.get_cell_param_value(i))
        # print(rd.get_cell_expect_value(i))
        print(rd.get_cell_url_value(i))














